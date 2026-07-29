from __future__ import annotations

import math
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ComparisonFieldMapping, DataComparisonSpec


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(value))).strip()


def decimal_value(
    value: Any,
    percentage: bool = False,
    currency: bool = False,
) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return Decimal(int(value))
    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return Decimal(str(value))
    text = clean_text(value).replace("\u00a0", "").replace(" ", "")
    has_percent = text.endswith("%")
    text = re.sub(r"[\u20ab$\u20ac\u00a3%]", "", text)
    if text.count(",") and text.count("."):
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") == 1 and len(text.rsplit(",", 1)[1]) <= 4:
        fraction = text.rsplit(",", 1)[1]
        text = text.replace(",", "" if currency and len(fraction) == 3 else ".")
    elif text.count(".") == 1 and currency and len(text.rsplit(".", 1)[1]) == 3:
        text = text.replace(".", "")
    else:
        text = text.replace(",", "")
    try:
        result = Decimal(text)
    except InvalidOperation:
        return None
    if percentage and has_percent:
        return result / Decimal(100)
    return result


def date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def normalized_value(value: Any, mapping: ComparisonFieldMapping) -> Any:
    if value is None or clean_text(value) == "":
        return ""
    if mapping.value_type in {"number", "currency", "percentage"}:
        numeric = decimal_value(
            value,
            mapping.value_type == "percentage",
            mapping.value_type == "currency",
        )
        if numeric is None:
            return clean_text(value)
        if mapping.displayed_precision is not None:
            quantizer = Decimal(1).scaleb(-mapping.displayed_precision)
            numeric = numeric.quantize(quantizer)
        return numeric
    if mapping.value_type == "date":
        return date_value(value)
    text = clean_text(value)
    return text if mapping.case_sensitive else text.casefold()


def values_match(expected: Any, actual: Any, mapping: ComparisonFieldMapping) -> bool:
    left = normalized_value(expected, mapping)
    right = normalized_value(actual, mapping)
    if isinstance(left, Decimal) and isinstance(right, Decimal):
        return abs(left - right) <= Decimal(str(mapping.tolerance))
    return left == right


def read_expected_rows(
    workbook_path: Path,
    sheet_name: str,
    header_row: int,
    key_columns: list[str],
    mappings: list[ComparisonFieldMapping],
) -> tuple[list[dict[str, Any]], list[str]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' was not found.")
        worksheet = workbook[sheet_name]
        headers = [clean_text(cell.value) for cell in worksheet[header_row]]
        header_positions = {
            header: index for index, header in enumerate(headers) if header
        }
        required = set(key_columns) | {item.excel_column for item in mappings}
        missing = sorted(required - set(header_positions))
        if missing:
            raise ValueError(f"Excel columns were not found: {', '.join(missing)}")
        rows: list[dict[str, Any]] = []
        warnings: list[str] = []
        seen_keys: set[tuple[str, ...]] = set()
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            row = {
                header: values[position] if position < len(values) else None
                for header, position in header_positions.items()
            }
            if not any(clean_text(value) for value in row.values()):
                continue
            key = tuple(clean_text(row.get(column)) for column in key_columns)
            if not any(key):
                warnings.append(
                    f"Row {len(rows) + header_row + 1} has an empty comparison key."
                )
            if key in seen_keys:
                warnings.append(f"Duplicate Excel key: {' | '.join(key)}")
            seen_keys.add(key)
            rows.append(row)
        return rows, warnings
    finally:
        workbook.close()


def compare_actual_rows(
    spec: DataComparisonSpec,
    actual_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    mapping_by_excel = {item.excel_column: item for item in spec.mappings}

    def expected_key(row: dict[str, Any]) -> tuple[str, ...]:
        return tuple(
            str(normalized_value(row.get(column), mapping_by_excel[column]))
            for column in spec.key_columns
        )

    key_ui_fields = {
        column: mapping_by_excel[column].ui_field
        for column in spec.key_columns
        if column in mapping_by_excel
    }
    if len(key_ui_fields) != len(spec.key_columns):
        missing = sorted(set(spec.key_columns) - set(key_ui_fields))
        raise ValueError(
            f"Every key column needs a UI mapping. Missing: {', '.join(missing)}"
        )

    def actual_key(row: dict[str, Any]) -> tuple[str, ...]:
        return tuple(
            str(normalized_value(
                row.get(key_ui_fields[column]),
                mapping_by_excel[column],
            ))
            for column in spec.key_columns
        )

    expected_index: dict[tuple[str, ...], dict[str, Any]] = {}
    duplicate_expected: set[tuple[str, ...]] = set()
    for row in spec.expected_rows:
        key = expected_key(row)
        if key in expected_index:
            duplicate_expected.add(key)
        expected_index[key] = row
    actual_index: dict[tuple[str, ...], dict[str, Any]] = {}
    duplicate_actual: set[tuple[str, ...]] = set()
    for row in actual_rows:
        key = actual_key(row)
        if key in actual_index:
            duplicate_actual.add(key)
        actual_index[key] = row

    results: list[dict[str, Any]] = []
    for key, expected in expected_index.items():
        if key in duplicate_expected or key in duplicate_actual:
            results.append({
                "key": list(key),
                "status": "needs_review",
                "differences": ["Duplicate key prevents a deterministic comparison."],
                "expected": expected,
                "actual": actual_index.get(key),
            })
            continue
        actual = actual_index.get(key)
        if actual is None:
            results.append({
                "key": list(key),
                "status": "missing_in_ui",
                "differences": [],
                "expected": expected,
                "actual": None,
            })
            continue
        differences = []
        for mapping in spec.mappings:
            expected_value = expected.get(mapping.excel_column)
            actual_value = actual.get(mapping.ui_field)
            if not values_match(expected_value, actual_value, mapping):
                differences.append({
                    "excel_column": mapping.excel_column,
                    "ui_field": mapping.ui_field,
                    "expected": expected_value,
                    "actual": actual_value,
                })
        results.append({
            "key": list(key),
            "status": "value_mismatch" if differences else "matched",
            "differences": differences,
            "expected": expected,
            "actual": actual,
        })
    for key, actual in actual_index.items():
        if key in expected_index:
            continue
        results.append({
            "key": list(key),
            "status": "extra_in_ui",
            "differences": [],
            "expected": None,
            "actual": actual,
        })

    summary: dict[str, int] = {}
    for result in results:
        summary[result["status"]] = summary.get(result["status"], 0) + 1
    return {
        "comparison_id": spec.id,
        "sheet": spec.sheet,
        "target_url": spec.target_url,
        "summary": summary,
        "rows": results,
    }
