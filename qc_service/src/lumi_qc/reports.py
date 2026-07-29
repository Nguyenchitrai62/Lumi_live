from __future__ import annotations

import html
import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import RunPlan
from .security import redact

LOG_HEADERS = [
    "Run ID",
    "Test Case",
    "Step",
    "Status",
    "Actual",
    "Expected",
    "Evidence",
    "URL",
    "Locator",
    "Confidence",
    "Retry Count",
    "Execution Mode",
    "Timing (ms)",
]


def safe_excel_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _safe_sheet(workbook, title: str):
    if title in workbook.sheetnames:
        del workbook[title]
    return workbook.create_sheet(title)


def generate_artifacts(
    plan: RunPlan,
    source_path: Path,
    results: dict[str, dict[str, Any]],
    output_dir: Path,
    comparison_results: list[dict[str, Any]] | None = None,
    bug_drafts: list[dict[str, Any]] | None = None,
) -> tuple[Path, Path]:
    stem = Path(plan.workbook_name).stem
    xlsx_path = output_dir / f"{stem}_executed_{plan.run_id}.xlsx"
    html_path = output_dir / f"{stem}_report_{plan.run_id}.html"
    shutil.copy2(source_path, xlsx_path)
    workbook = load_workbook(xlsx_path, data_only=False, read_only=False)
    log_sheet = _safe_sheet(workbook, "Agent_Run_Log")
    result_sheet = _safe_sheet(workbook, "Agent_Result")
    log_sheet.append(LOG_HEADERS)
    result_sheet.append(["Run ID", "Test Case", "Step", "Instruction", "Status", "Actual", "Evidence"])

    rows: list[dict[str, Any]] = []
    for case in plan.test_cases:
        for step in case.steps:
            result = results.get(step.id, {})
            status = result.get("status", step.status.value)
            actual = result.get("actual", "")
            evidence = result.get("evidence", "")
            expected = result.get("expected") or step.expected
            writeback = step.writeback
            mapped = False
            source_sheet = workbook[writeback.sheet]
            if writeback.status_cell:
                source_sheet[writeback.status_cell] = safe_excel_text(status)
                mapped = True
            if writeback.actual_cell:
                source_sheet[writeback.actual_cell] = safe_excel_text(actual)
                mapped = True
            if writeback.evidence_cell:
                source_sheet[writeback.evidence_cell] = safe_excel_text(evidence)
                mapped = True
            log_sheet.append(
                [
                    plan.run_id,
                    case.id,
                    step.id,
                    status,
                    safe_excel_text(actual),
                    safe_excel_text(expected),
                    safe_excel_text(evidence),
                    result.get("url", ""),
                    result.get("locator", ""),
                    result.get("confidence"),
                    result.get("retry_count", 0),
                    result.get("execution_mode", "step"),
                    safe_excel_text(json.dumps(result.get("timings", {}), ensure_ascii=False)),
                ]
            )
            if not mapped:
                result_sheet.append(
                    [
                        plan.run_id,
                        case.id,
                        step.id,
                        step.instruction,
                        status,
                        safe_excel_text(actual),
                        safe_excel_text(evidence),
                    ]
                )
            rows.append(
                {
                    "case": case.id,
                    "title": case.title,
                    "step": step.id,
                    "instruction": step.instruction,
                    "status": status,
                    "actual": actual,
                    "expected": expected,
                    "evidence": evidence,
                    "url": result.get("url", ""),
                }
            )

    log_sheet.freeze_panes = "A2"
    result_sheet.freeze_panes = "A2"
    comparison_rows = []
    if comparison_results:
        comparison_sheet = _safe_sheet(workbook, "Agent_Data_Comparison")
        comparison_sheet.append(
            [
                "Comparison ID",
                "Sheet",
                "Key",
                "Status",
                "Differences",
                "Expected",
                "Actual",
            ]
        )
        for comparison in comparison_results:
            for comparison_row in comparison.get("rows", []):
                comparison_sheet.append(
                    [
                        comparison.get("comparison_id", ""),
                        comparison.get("sheet", ""),
                        safe_excel_text(" | ".join(comparison_row.get("key", []))),
                        comparison_row.get("status", ""),
                        safe_excel_text(
                            json.dumps(
                                comparison_row.get("differences", []),
                                ensure_ascii=False,
                                default=str,
                            )
                        ),
                        safe_excel_text(
                            json.dumps(
                                comparison_row.get("expected"),
                                ensure_ascii=False,
                                default=str,
                            )
                        ),
                        safe_excel_text(
                            json.dumps(
                                comparison_row.get("actual"),
                                ensure_ascii=False,
                                default=str,
                            )
                        ),
                    ]
                )
                comparison_rows.append(comparison_row)
        comparison_sheet.freeze_panes = "A2"
    for source_sheet in workbook.worksheets:
        for row in source_sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or cell.value.startswith("="):
                    continue
                redacted = str(redact(cell.value))
                if redacted != cell.value:
                    cell.value = redacted
    workbook.save(xlsx_path)
    workbook.close()

    summary: dict[str, int] = {}
    for row in rows:
        summary[row["status"]] = summary.get(row["status"], 0) + 1
    summary_cards = "".join(
        f"<li><strong>{html.escape(status)}</strong><span>{count}</span></li>"
        for status, count in sorted(summary.items())
    )
    timeline = "".join(
        f"""
        <article class="step status-{html.escape(str(row['status']))}">
          <header><strong>{html.escape(str(row['case']))} · {html.escape(str(row['step']))}</strong>
          <span>{html.escape(str(row['status']))}</span></header>
          <p>{html.escape(str(row['instruction']))}</p>
          <dl>
            <dt>Expected</dt><dd>{html.escape(str(row['expected']))}</dd>
            <dt>Actual</dt><dd>{html.escape(str(row['actual']))}</dd>
            <dt>Evidence</dt><dd>{html.escape(str(row['evidence']))}</dd>
            <dt>URL</dt><dd>{html.escape(str(row['url']))}</dd>
          </dl>
        </article>
        """
        for row in rows
    )
    comparison_timeline = "".join(
        f"""
        <article class="step status-{html.escape(str(row.get('status', '')))}">
          <header><strong>Data comparison Â· {html.escape(' | '.join(row.get('key', [])))}</strong>
          <span>{html.escape(str(row.get('status', '')))}</span></header>
          <dl>
            <dt>Differences</dt><dd>{html.escape(json.dumps(row.get('differences', []), ensure_ascii=False, default=str))}</dd>
            <dt>Expected</dt><dd>{html.escape(json.dumps(row.get('expected'), ensure_ascii=False, default=str))}</dd>
            <dt>Actual</dt><dd>{html.escape(json.dumps(row.get('actual'), ensure_ascii=False, default=str))}</dd>
          </dl>
        </article>
        """
        for row in comparison_rows
    )
    bug_timeline = "".join(
        f"""
        <article class="step">
          <header><strong>Redmine draft Â· {html.escape(str(item.get('id', '')))}</strong>
          <span>{html.escape(str(item.get('status', 'draft')))}</span></header>
          <p>{html.escape(str(item.get('draft', {}).get('subject', '')))}</p>
          <p>Issue: {html.escape(str(item.get('submitted_issue_url', '') or 'not submitted'))}</p>
        </article>
        """
        for item in (bug_drafts or [])
    )
    report = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width">
<title>Lumi QC report · {html.escape(plan.run_id)}</title>
<style>
body{{font:14px/1.5 system-ui,sans-serif;margin:0;background:#f6f3fa;color:#241b2f}}
main{{max-width:1080px;margin:auto;padding:32px}}h1{{margin-bottom:4px}}
.summary{{display:flex;gap:12px;list-style:none;padding:0;flex-wrap:wrap}}
.summary li{{background:white;border:1px solid #ddd2e8;border-radius:12px;padding:12px 16px;display:grid}}
.summary span{{font-size:24px}}.step{{background:white;border-left:5px solid #a99ab9;margin:12px 0;padding:16px;border-radius:10px}}
.status-passed{{border-left-color:#279b68}}.status-failed_product{{border-left-color:#d84b4b}}
.status-spec_issue,.status-needs_review{{border-left-color:#d89b26}}header{{display:flex;justify-content:space-between;gap:20px}}
dl{{display:grid;grid-template-columns:100px 1fr;gap:5px 12px}}dt{{font-weight:700}}dd{{margin:0;white-space:pre-wrap}}
</style></head><body><main>
<h1>Lumi QC execution report</h1>
<p>Workbook: {html.escape(plan.workbook_name)} · Run: {html.escape(plan.run_id)}</p>
<ul class="summary">{summary_cards}</ul>
<section>{timeline}</section>
<section>{comparison_timeline}</section>
<section>{bug_timeline}</section>
<script type="application/json" id="lumi-qc-plan">{html.escape(json.dumps(plan.model_dump(mode="json"), ensure_ascii=False))}</script>
</main></body></html>"""
    html_path.write_text(report, encoding="utf-8")
    return xlsx_path, html_path
