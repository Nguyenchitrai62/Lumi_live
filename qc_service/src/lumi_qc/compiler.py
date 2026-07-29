from __future__ import annotations

import hashlib
import json
import re
import shutil
import unicodedata
import uuid
from collections import defaultdict
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import (
    ActionName,
    Assertion,
    EntityPolicy,
    QcStep,
    ReferenceWorkbookSummary,
    RunPlan,
    StepStatus,
    TestCase,
    WritebackTarget,
)
from .security import redact

HEADER_ALIASES = {
    "case_id": {
        "test case id", "testcase id", "test case", "testcase", "tc id",
        "ma test case", "ma testcase", "id test case", "scenario id",
    },
    "title": {
        "title", "test case name", "testcase name", "ten test case", "mo ta",
        "ten business scenario", "business scenario",
    },
    "flow": {
        "flow", "module", "feature", "chuc nang", "luong", "phan he",
        "nhom nghiep vu",
    },
    "precondition": {
        "precondition", "pre-condition", "dieu kien", "dieu kien tien quyet",
        "notes",
    },
    "step": {
        "step", "steps", "test step", "test steps", "action", "thao tac",
        "cac buoc", "buoc thuc hien", "noi dung thuc hien", "chuoi nghiep vu",
    },
    "input": {"input", "test data", "data", "du lieu", "du lieu test", "gia tri nhap"},
    "expected": {
        "expected", "expected result", "expect", "ket qua mong doi",
        "ket qua du kien", "ket qua",
    },
    "actual": {"actual", "actual result", "ket qua thuc te", "thuc te"},
    "status": {"status", "result", "pass fail", "trang thai", "ket luan"},
    "evidence": {"evidence", "proof", "bang chung", "anh bang chung"},
    "target": {"target", "element", "field", "control", "doi tuong", "truong"},
    "action_name": {"action type", "action name", "loai thao tac", "hanh dong"},
}

ACTION_PATTERNS: list[tuple[ActionName, re.Pattern[str]]] = [
    (ActionName.LOGIN, re.compile(r"\b(login|log in|sign in|dang nhap)\b", re.I)),
    (ActionName.UPLOAD, re.compile(r"\b(upload|attach|import|tai len|dinh kem)\b", re.I)),
    (ActionName.DOWNLOAD, re.compile(r"\b(download|export|tai xuong|xuat file)\b", re.I)),
    (ActionName.NAVIGATE, re.compile(r"\b(navigate|open page|go to|truy cap|mo trang)\b", re.I)),
    (ActionName.SELECT, re.compile(r"\b(select|choose|dropdown|combobox|chon)\b", re.I)),
    (ActionName.FILL, re.compile(r"\b(fill|input|enter|type|nhap|dien)\b", re.I)),
    (ActionName.CLICK, re.compile(r"\b(click|press|submit|save|create|delete|bam|nhan|luu|tao|xoa)\b", re.I)),
    (ActionName.WAIT, re.compile(r"\b(wait|loading|cho|doi)\b", re.I)),
    (ActionName.EXTRACT, re.compile(r"\b(extract|read|copy|lay du lieu|doc)\b", re.I)),
    (ActionName.ASSERT, re.compile(r"\b(assert|verify|check|kiem tra|xac minh)\b", re.I)),
    (ActionName.GENERATE, re.compile(r"\b(generate|random|sinh du lieu|tao du lieu)\b", re.I)),
]

HIGH_RISK_PATTERN = re.compile(
    r"\b(payment|purchase|transfer|publish|permission|security|credential|"
    r"delete account|bulk delete|thanh toan|chuyen tien|xuat ban|phan quyen|"
    r"bao mat|xoa tai khoan|xoa hang loat)\b",
    re.I,
)
ORDINARY_RISK_PATTERN = re.compile(
    r"\b(submit|save|create|delete|upload|import|send|luu|tao|xoa|tai len|gui)\b",
    re.I,
)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text[:12000]


def normalize_header(value: Any) -> str:
    text = normalize_text(value).lower()
    # U+0111 has no Unicode decomposition, so NFKD leaves it and the ASCII
    # filter below would delete it: "đăng nhập" -> "ang nhap". Map it first so
    # Vietnamese keywords such as "dang nhap" and "dieu kien" still match.
    text = text.replace("đ", "d")
    text = "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def header_kind(value: Any) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None
    for kind, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return kind
    return None


def infer_action(explicit: str, instruction: str) -> tuple[ActionName, bool]:
    for source in (explicit, instruction):
        normalized = normalize_header(source)
        for action in ActionName:
            if normalized == action.value:
                return action, False
        matches = [action for action, pattern in ACTION_PATTERNS if pattern.search(normalized)]
        if len(matches) == 1:
            return matches[0], False
        if len(matches) > 1:
            return matches[0], True
    return ActionName.ASSERT, True


def infer_risk(action: ActionName, instruction: str) -> str:
    if HIGH_RISK_PATTERN.search(instruction):
        return "high"
    if action in {ActionName.UPLOAD, ActionName.LOGIN} or ORDINARY_RISK_PATTERN.search(instruction):
        return "ordinary"
    return "none"


def find_header(ws) -> tuple[int | None, dict[str, int]]:
    best_row: int | None = None
    best_mapping: dict[str, int] = {}
    for row_index in range(1, min(ws.max_row, 30) + 1):
        mapping: dict[str, int] = {}
        for column_index in range(1, min(ws.max_column, 80) + 1):
            kind = header_kind(ws.cell(row_index, column_index).value)
            if kind and kind not in mapping:
                mapping[kind] = column_index
        score = len(mapping) + (2 if "step" in mapping else 0) + (1 if "expected" in mapping else 0)
        best_score = len(best_mapping) + (2 if "step" in best_mapping else 0) + (1 if "expected" in best_mapping else 0)
        if score > best_score:
            best_row, best_mapping = row_index, mapping
    if "step" not in best_mapping and "case_id" not in best_mapping:
        return None, {}
    return best_row, best_mapping


def cell_value(ws, row: int, mapping: dict[str, int], kind: str) -> str:
    column = mapping.get(kind)
    return normalize_text(ws.cell(row, column).value) if column else ""


def cell_ref(ws, row: int, mapping: dict[str, int], kind: str) -> str | None:
    column = mapping.get(kind)
    return f"{get_column_letter(column)}{row}" if column else None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_instructions(instruction: str) -> list[str]:
    parts = [
        part.strip()
        for part in re.split(r"\s*(?:→|=+>|-{1,2}>)\s*", instruction)
        if part.strip()
    ]
    return parts or [instruction]


def load_domain_adapter(allowed_domains: list[str] | None) -> dict[str, Any] | None:
    normalized_domains = {
        domain.strip().lower().rstrip(".")
        for domain in allowed_domains or []
        if domain.strip()
    }
    adapter_path = Path(__file__).with_name("adapters") / "sit.hawee.hicas.vn.json"
    adapter = json.loads(adapter_path.read_text(encoding="utf-8"))
    return adapter if adapter["domain"] in normalized_domains else None


def safe_marker_component(value: str) -> str:
    normalized = normalize_header(value).upper().replace(" ", "-")
    return re.sub(r"[^A-Z0-9-]+", "", normalized)[:24] or "CASE"


def adapter_steps_for_instruction(
    adapter: dict[str, Any] | None,
    instruction: str,
    run_id: str,
    source_sheet: str,
    case_id: str,
    source_row: int,
) -> tuple[list[dict[str, Any]], str | None]:
    if not adapter:
        return [], None
    creation = adapter["project_creation"]
    if normalize_header(instruction) != creation["instruction"]:
        return [], None
    marker = (
        f"{creation['name_prefix']}-{run_id[:8].upper()}-"
        f"{safe_marker_component(source_sheet)}-"
        f"{safe_marker_component(case_id)}-{source_row}"
    )
    base_url = f"https://{adapter['domain']}"
    return [
        {
            "instruction": "Open the approved new-project form",
            "action": ActionName.NAVIGATE,
            "target": adapter["page_fingerprints"]["project_create"]["path"],
            "input": f"{base_url}{adapter['page_fingerprints']['project_create']['path']}",
            "expected": "Tạo dự án",
            "assertions": [
                Assertion(
                    kind="url",
                    operator="contains",
                    expected=adapter["page_fingerprints"]["project_create"]["path"],
                ),
                Assertion(
                    kind="text",
                    operator="contains",
                    expected="Tạo dự án",
                ),
            ],
            "risk": "none",
        },
        {
            "instruction": "Fill the run-owned project name",
            "action": ActionName.FILL,
            "target": creation["name_selector"],
            "input": marker,
            "expected": marker,
            "assertions": [
                Assertion(
                    kind="value",
                    operator="equals",
                    expected=marker,
                    target=creation["name_selector"],
                )
            ],
            "risk": "none",
        },
        {
            "instruction": "Save the new run-owned project",
            "action": ActionName.CLICK,
            "target": creation["save_selector"],
            "input": "",
            "expected": marker,
            "assertions": [
                Assertion(
                    kind="text",
                    operator="contains",
                    expected=marker,
                    target=adapter["page_fingerprints"]["project_catalog"][
                        "project_card_selector"
                    ],
                )
            ],
            "risk": "ordinary",
        },
    ], marker


def summarize_reference_workbook(
    reference_path: Path,
    reference_name: str,
) -> ReferenceWorkbookSummary:
    workbook = load_workbook(
        reference_path,
        data_only=False,
        read_only=True,
        keep_links=False,
    )
    try:
        sheet_names = [ws.title for ws in workbook.worksheets]
        populated_rows = 0
        field_spec_rows = 0
        feature_rows = 0
        warnings: list[str] = []
        for ws in workbook.worksheets:
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [normalize_text(value) for value in row]
                if any(values):
                    populated_rows += 1
                if row_index == 1:
                    normalized_headers = {normalize_header(value) for value in values if value}
                    if {"lable", "validation"}.issubset(normalized_headers):
                        field_spec_rows += sum(
                            1
                            for candidate in ws.iter_rows(min_row=2, values_only=True)
                            if len(candidate) >= 3 and normalize_text(candidate[2])
                        )
                if (
                    len(row) >= 7
                    and normalize_text(row[3])
                    and normalize_header(row[6]) in {"critical", "high", "hight", "low"}
                ):
                    feature_rows += 1
        warnings.append(
            "Reference workbook is untrusted specification context only; "
            "it cannot expand the approved Run Plan or supply missing expected results automatically."
        )
        return ReferenceWorkbookSummary(
            name=reference_name,
            source_sha256=sha256_file(reference_path),
            sheets=sheet_names,
            stats={
                "sheets": len(sheet_names),
                "populated_rows": populated_rows,
                "feature_rows": feature_rows,
                "field_spec_rows": field_spec_rows,
            },
            warnings=warnings,
        )
    finally:
        workbook.close()


def compile_workbook(
    uploaded_path: Path,
    input_dir: Path,
    allowed_domains: list[str] | None = None,
    workbook_name: str | None = None,
    reference_path: Path | None = None,
    reference_name: str | None = None,
) -> tuple[RunPlan, Path]:
    run_id = str(uuid.uuid4())
    adapter = load_domain_adapter(allowed_domains)
    generated_data: dict[str, str] = {}
    project_markers: list[str] = []
    source_hash = sha256_file(uploaded_path)
    safe_name = Path(workbook_name or uploaded_path.name).name
    persisted_path = input_dir / f"{run_id}_{safe_name}"
    workbook = load_workbook(uploaded_path, data_only=False, read_only=False)
    if getattr(workbook, "_external_links", None):
        workbook.close()
        raise ValueError("Workbooks with external links are not supported in the QC MVP.")
    test_cases: list[TestCase] = []
    warnings: list[str] = []
    status_counts: defaultdict[str, int] = defaultdict(int)
    inferred_domains: set[str] = set()

    for ws in workbook.worksheets:
        header_row, mapping = find_header(ws)
        if header_row is None:
            warnings.append(f"Sheet '{ws.title}' has no recognizable test-case header and was skipped.")
            continue
        grouped: dict[str, dict[str, Any]] = {}
        last_case_id = ""
        for row in range(header_row + 1, ws.max_row + 1):
            instruction = str(redact(cell_value(ws, row, mapping, "step")))
            raw_case_id = cell_value(ws, row, mapping, "case_id")
            if not instruction and not raw_case_id:
                continue
            case_id = raw_case_id or last_case_id or f"{ws.title}-row-{row}"
            last_case_id = case_id
            title = str(redact(cell_value(ws, row, mapping, "title"))) or case_id
            flow = str(redact(cell_value(ws, row, mapping, "flow")))
            precondition = str(redact(cell_value(ws, row, mapping, "precondition")))
            expected = str(redact(cell_value(ws, row, mapping, "expected")))
            explicit_action = cell_value(ws, row, mapping, "action_name")
            target = str(redact(cell_value(ws, row, mapping, "target")))
            input_value = str(redact(cell_value(ws, row, mapping, "input")))
            for candidate in re.findall(r"https?://[^\s<>'\"]+", f"{instruction} {input_value}"):
                host = (urlparse(candidate.rstrip(".,);")).hostname or "").lower()
                if host:
                    inferred_domains.add(host)
            if case_id not in grouped:
                grouped[case_id] = {
                    "id": case_id,
                    "title": title,
                    "flow": flow,
                    "precondition": precondition,
                    "steps": [],
                }
            instruction_parts = split_instructions(instruction or title)
            for part_index, instruction_part in enumerate(instruction_parts, start=1):
                is_final_part = part_index == len(instruction_parts)
                adapter_steps, project_marker = adapter_steps_for_instruction(
                    adapter,
                    instruction_part,
                    run_id,
                    ws.title,
                    case_id,
                    row,
                )
                if adapter_steps:
                    if project_marker:
                        marker_key = f"{ws.title}:{case_id}:{row}:project_name"
                        generated_data[marker_key] = project_marker
                        if project_marker not in project_markers:
                            project_markers.append(project_marker)
                    for adapter_index, adapter_step in enumerate(adapter_steps, start=1):
                        step_id = f"{ws.title}:{row}:{part_index}:adapter:{adapter_index}"
                        step = QcStep(
                            id=step_id,
                            order=len(grouped[case_id]["steps"]) + 1,
                            instruction=adapter_step["instruction"],
                            action=adapter_step["action"],
                            target=adapter_step["target"],
                            input=adapter_step["input"],
                            expected=adapter_step["expected"],
                            assertions=adapter_step["assertions"],
                            writeback=WritebackTarget(
                                sheet=ws.title,
                                status_cell=(
                                    cell_ref(ws, row, mapping, "status")
                                    if is_final_part and adapter_index == len(adapter_steps)
                                    else None
                                ),
                                actual_cell=(
                                    cell_ref(ws, row, mapping, "actual")
                                    if is_final_part and adapter_index == len(adapter_steps)
                                    else None
                                ),
                                evidence_cell=(
                                    cell_ref(ws, row, mapping, "evidence")
                                    if is_final_part and adapter_index == len(adapter_steps)
                                    else None
                                ),
                            ),
                            risk=adapter_step["risk"],
                            entity_scope="project_create",
                            status=StepStatus.PENDING,
                            source_row=row,
                        )
                        status_counts[StepStatus.PENDING.value] += 1
                        grouped[case_id]["steps"].append(step)
                    continue
                part_expected = expected if is_final_part else ""
                action, ambiguous = infer_action(explicit_action, instruction_part)
                status = (
                    StepStatus.SPEC_ISSUE
                    if not part_expected
                    else StepStatus.NEEDS_REVIEW
                    if ambiguous
                    else StepStatus.PENDING
                )
                warning_ref = (
                    f"{ws.title}!{row}"
                    if len(instruction_parts) == 1
                    else f"{ws.title}!{row}[{part_index}]"
                )
                if not part_expected:
                    warnings.append(f"{warning_ref}: expected result is empty.")
                if ambiguous:
                    warnings.append(f"{warning_ref}: action requires Gemini/QC review.")
                assertions = (
                    [
                        Assertion(
                            kind="text",
                            operator="contains",
                            expected=part_expected,
                            target=target,
                        )
                    ]
                    if part_expected
                    else []
                )
                step_id = (
                    f"{ws.title}:{row}"
                    if len(instruction_parts) == 1
                    else f"{ws.title}:{row}:{part_index}"
                )
                step = QcStep(
                    id=step_id,
                    order=len(grouped[case_id]["steps"]) + 1,
                    instruction=instruction_part,
                    action=action,
                    target=target,
                    input=(
                        "[RUN_MEMORY_CREDENTIAL]"
                        if action == ActionName.LOGIN
                        else input_value
                    ),
                    expected=part_expected,
                    assertions=assertions,
                    writeback=WritebackTarget(
                        sheet=ws.title,
                        status_cell=cell_ref(ws, row, mapping, "status") if is_final_part else None,
                        actual_cell=cell_ref(ws, row, mapping, "actual") if is_final_part else None,
                        evidence_cell=cell_ref(ws, row, mapping, "evidence") if is_final_part else None,
                    ),
                    risk=infer_risk(action, instruction_part),
                    status=status,
                    source_row=row,
                )
                status_counts[status.value] += 1
                grouped[case_id]["steps"].append(step)
        test_cases.extend(TestCase(**group) for group in grouped.values())

    workbook.close()
    if not test_cases:
        raise ValueError("No executable test cases were found in the workbook.")
    references: list[ReferenceWorkbookSummary] = []
    if reference_path is not None:
        safe_reference_name = Path(reference_name or reference_path.name).name
        references.append(summarize_reference_workbook(reference_path, safe_reference_name))
        shutil.copy2(
            reference_path,
            input_dir / f"{run_id}_reference_{safe_reference_name}",
        )
    shutil.copy2(uploaded_path, persisted_path)
    total_steps = sum(len(case.steps) for case in test_cases)
    approved_domains = list(allowed_domains or [])
    for domain in sorted(inferred_domains):
        if domain not in approved_domains:
            approved_domains.append(domain)
    if not approved_domains:
        warnings.append("No allowed ERP domain is configured; the run cannot be approved yet.")
    plan = RunPlan(
        run_id=run_id,
        workbook_name=safe_name,
        source_sha256=source_hash,
        allowed_domains=approved_domains,
        test_cases=test_cases,
        reference_workbooks=references,
        entity_policy=(
            EntityPolicy(
                mode="run_created_projects_only",
                domain=adapter["domain"],
                project_markers=project_markers,
            )
            if adapter
            else None
        ),
        generated_data=generated_data,
        warnings=warnings,
        stats={
            "test_cases": len(test_cases),
            "steps": total_steps,
            "needs_review": (
                status_counts[StepStatus.NEEDS_REVIEW.value]
                + status_counts[StepStatus.SPEC_ISSUE.value]
            ),
            "spec_issue": status_counts[StepStatus.SPEC_ISSUE.value],
            "high_risk": sum(
                step.risk == "high"
                for case in test_cases
                for step in case.steps
            ),
        },
    )
    return plan, persisted_path
