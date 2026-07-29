from __future__ import annotations

import base64
import hashlib
import sqlite3
import tempfile
import unittest
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from lumi_qc.app import create_app
from lumi_qc.comparison import compare_actual_rows
from lumi_qc.models import ComparisonFieldMapping, DataComparisonSpec
from lumi_qc.repository import QcRepository


class Qc020Test(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.token = "qc-020-test-installation-token-with-entropy"
        self.app = create_app(self.root / "data", self.token)
        self.client = TestClient(self.app)
        self.headers = {
            "Origin": "chrome-extension://abcdefghijklmnop",
            "X-Lumi-Installation-Token": self.token,
        }

    def tearDown(self) -> None:
        self.client.close()
        self.temporary.cleanup()

    def create_prompt_run(self, *, expected: str = "The page is visible") -> dict:
        provider_key = "AI" + "za" + "SyABCDEFGHIJKLMNOPQRSTUVWXYZ123456"
        example_password = "ExampleOnly!42"
        response = self.client.post(
            "/v1/runs/from-prompt",
            headers=self.headers,
            json={
                "prompt": (
                    f"tk: admin/{example_password} "
                    f"key={provider_key}"
                ),
                "title": "Credential-safe prompt",
                "target_url": "https://sit.hawee.hicas.vn/tong-quan?tab=du-an",
                "target_fingerprint": "",
                "knowledge_version": "0.2.0",
                "allowed_domains": ["sit.hawee.hicas.vn"],
                "execution_mode": "step",
                "steps": [
                    {
                        "instruction": "Login with password=do-not-persist",
                        "action": "login",
                        "target": "Login form",
                        "input": f"admin/{example_password}",
                        "expected": expected,
                        "assertions": [],
                        "risk": "ordinary",
                        "entity_scope": "none",
                    }
                ],
            },
        )
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()

    def finish_run(self, run: dict, status: str = "passed", retry_count: int = 0) -> dict:
        run_id = run["run_id"]
        step_id = run["plan"]["test_cases"][0]["steps"][0]["id"]
        approved = self.client.post(f"/v1/runs/{run_id}/approve", headers=self.headers)
        self.assertEqual(approved.status_code, 200, approved.text)
        self.assertEqual(
            self.client.post(f"/v1/runs/{run_id}/start", headers=self.headers).status_code,
            200,
        )
        self.assertEqual(
            self.client.post(
                f"/v1/runs/{run_id}/steps/{step_id}/begin",
                headers=self.headers,
            ).status_code,
            200,
        )
        recorded = self.client.post(
            f"/v1/runs/{run_id}/steps/{step_id}/record",
            headers=self.headers,
            json={
                "status": status,
                "actual": "Actual",
                "expected": "Expected",
                "evidence": "Verified after retry",
                "url": "https://sit.hawee.hicas.vn/tong-quan",
                "retry_count": retry_count,
            },
        )
        self.assertEqual(recorded.status_code, 200, recorded.text)
        completed = self.client.post(
            f"/v1/runs/{run_id}/complete",
            headers=self.headers,
            json={
                "status": "completed" if status == "passed" else "failed",
                "summary": "terminal",
            },
        )
        self.assertEqual(completed.status_code, 200, completed.text)
        return completed.json()

    def test_prompt_credentials_are_redacted_and_schedule_clones_fresh_marker(self) -> None:
        run = self.create_prompt_run()
        serialized = str(run)
        self.assertNotIn("ExampleOnly", serialized)
        self.assertNotIn("AIzaSy", serialized)
        step = run["plan"]["test_cases"][0]["steps"][0]
        self.assertEqual(step["input"], "[RUN_MEMORY_CREDENTIAL]")
        source_path = self.app.state.repository.source_path(run["run_id"])
        source = load_workbook(source_path, data_only=False)
        prompt_text = str(source["Prompt_Run"]["A2"].value)
        source.close()
        self.assertNotIn("ExampleOnly", prompt_text)
        self.assertNotIn("AIzaSy", prompt_text)

        completed = self.finish_run(run)
        schedule = self.client.post(
            "/v1/schedules",
            headers=self.headers,
            json={
                "name": "Weekday smoke",
                "template_run_id": completed["run_id"],
                "local_time": "08:15",
                "days_of_week": [0, 1, 2, 3, 4],
                "timezone": "Asia/Bangkok",
                "enabled": True,
            },
        )
        self.assertEqual(schedule.status_code, 200, schedule.text)
        cloned = self.client.post(
            f"/v1/schedules/{schedule.json()['id']}/run-now",
            headers=self.headers,
        )
        self.assertEqual(cloned.status_code, 200, cloned.text)
        clone = cloned.json()["run"]
        self.assertEqual(clone["status"], "running")
        self.assertNotEqual(clone["run_id"], completed["run_id"])
        self.assertNotEqual(
            clone["plan"]["generated_data"]["run_marker"],
            completed["plan"]["generated_data"]["run_marker"],
        )

    def test_bug_draft_requires_reproducible_product_failure_and_evidence_scope(self) -> None:
        run = self.create_prompt_run()
        run_id = run["run_id"]
        step_id = run["plan"]["test_cases"][0]["steps"][0]["id"]
        denied = self.client.post(
            f"/v1/runs/{run_id}/bug-drafts",
            headers=self.headers,
            json={
                "step_id": step_id,
                "module": "overview",
                "subject": "[LUMI][HICAS][overview] test – failure",
                "description": "Expected/actual evidence",
                "classification": "failed_product",
                "expected": "Expected",
                "actual": "Actual",
                "evidence": "One observation",
            },
        )
        self.assertEqual(denied.status_code, 409)
        self.finish_run(run, status="failed_product", retry_count=2)
        created = self.client.post(
            f"/v1/runs/{run_id}/bug-drafts",
            headers=self.headers,
            json={
                "step_id": step_id,
                "module": "overview",
                "subject": "[LUMI][HICAS][overview] test – failure",
                "description": "Expected/actual evidence",
                "classification": "failed_product",
                "expected": "Expected",
                "actual": "Actual",
                "evidence": "Reproduced twice",
                "confidence": 0.9,
            },
        )
        self.assertEqual(created.status_code, 200, created.text)
        self.assertEqual(created.json()["status"], "draft")

        invalid_scope = self.client.post(
            f"/v1/runs/{run_id}/evidence",
            headers=self.headers,
            json={
                "step_id": step_id,
                "mime_type": "image/jpeg",
                "data_base64": base64.b64encode(b"jpeg").decode(),
                "scope": "existing_project",
                "url": "https://sit.hawee.hicas.vn/",
            },
        )
        self.assertEqual(invalid_scope.status_code, 422)
        evidence = self.client.post(
            f"/v1/runs/{run_id}/evidence",
            headers=self.headers,
            json={
                "step_id": step_id,
                "mime_type": "image/jpeg",
                "data_base64": base64.b64encode(b"jpeg").decode(),
                "scope": "owned_sandbox",
                "url": "https://sit.hawee.hicas.vn/",
            },
        )
        self.assertEqual(evidence.status_code, 200, evidence.text)
        self.assertTrue(Path(evidence.json()["path"]).exists())

    def test_comparison_normalizes_locale_and_classifies_all_mismatch_types(self) -> None:
        mappings = [
            ComparisonFieldMapping(
                excel_column="Code",
                ui_field="Mã",
                value_type="text",
                case_sensitive=False,
            ),
            ComparisonFieldMapping(
                excel_column="Amount",
                ui_field="Giá trị",
                value_type="currency",
            ),
            ComparisonFieldMapping(
                excel_column="Percent",
                ui_field="Tỷ lệ",
                value_type="percentage",
            ),
            ComparisonFieldMapping(
                excel_column="Date",
                ui_field="Ngày",
                value_type="date",
            ),
        ]
        spec = DataComparisonSpec(
            id="compare-1",
            sheet="Data",
            header_row=1,
            key_columns=["Code"],
            mappings=mappings,
            target_url="https://sit.hawee.hicas.vn/grid",
            expected_rows=[
                {"Code": " A ", "Amount": 1234, "Percent": 0.1, "Date": "2026-07-29"},
                {"Code": "B", "Amount": 2000, "Percent": 0.2, "Date": "2026-07-30"},
                {"Code": "C", "Amount": 3000, "Percent": 0.3, "Date": "2026-07-31"},
            ],
        )
        result = compare_actual_rows(
            spec,
            [
                {"Mã": "a", "Giá trị": "1.234 ₫", "Tỷ lệ": "10%", "Ngày": "29/07/2026"},
                {"Mã": "B", "Giá trị": "2.500 ₫", "Tỷ lệ": "20%", "Ngày": "30/07/2026"},
                {"Mã": "D", "Giá trị": "4.000 ₫", "Tỷ lệ": "40%", "Ngày": "01/08/2026"},
            ],
        )
        self.assertEqual(result["summary"]["matched"], 1)
        self.assertEqual(result["summary"]["value_mismatch"], 1)
        self.assertEqual(result["summary"]["missing_in_ui"], 1)
        self.assertEqual(result["summary"]["extra_in_ui"], 1)

        duplicate = spec.model_copy(
            update={"expected_rows": [{"Code": "X"}, {"Code": "X"}]}
        )
        duplicate_result = compare_actual_rows(duplicate, [{"Mã": "X"}])
        self.assertEqual(duplicate_result["summary"]["needs_review"], 1)

    def test_comparison_api_requires_complete_pagination_and_writes_result_sheet(self) -> None:
        workbook_path = self.root / "comparison.xlsx"
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append(["Code", "Amount"])
        data.append(["A", 1000])
        data.append(["B", 2000])
        notes = workbook.create_sheet("Notes")
        notes.merge_cells("A1:B1")
        notes["A1"] = "Preserved"
        notes["D1"] = "=SUM(1,2)"
        workbook.save(workbook_path)
        workbook.close()
        source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
        compiled = self.client.post(
            "/v1/comparisons/compile",
            headers=self.headers,
            files={
                "file": (
                    "comparison.xlsx",
                    workbook_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={
                "sheet": "Data",
                "header_row": "1",
                "key_columns": '["Code"]',
                "mappings": (
                    '[{"excel_column":"Code","ui_field":"Mã","value_type":"text"},'
                    '{"excel_column":"Amount","ui_field":"Giá trị","value_type":"currency"}]'
                ),
                "target_url": "https://sit.hawee.hicas.vn/grid",
                "allowed_domains": '["sit.hawee.hicas.vn"]',
                "knowledge_version": "0.2.0",
                "target_fingerprint": "",
                "execution_mode": "step",
            },
        )
        self.assertEqual(compiled.status_code, 200, compiled.text)
        run = compiled.json()
        run_id = run["run_id"]
        comparison_id = run["plan"]["comparison_specs"][0]["id"]
        step_id = run["plan"]["test_cases"][0]["steps"][0]["id"]
        before_begin = self.client.post(
            f"/v1/runs/{run_id}/comparisons/{comparison_id}/actual",
            headers=self.headers,
            json={"rows": [], "complete": True, "page_count": 1},
        )
        self.assertEqual(before_begin.status_code, 409)
        self.client.post(f"/v1/runs/{run_id}/approve", headers=self.headers)
        self.client.post(f"/v1/runs/{run_id}/start", headers=self.headers)
        self.client.post(
            f"/v1/runs/{run_id}/steps/{step_id}/begin",
            headers=self.headers,
        )
        incomplete = self.client.post(
            f"/v1/runs/{run_id}/comparisons/{comparison_id}/actual",
            headers=self.headers,
            json={
                "rows": [{"Mã": "A", "Giá trị": "1.000 ₫"}],
                "complete": False,
                "page_count": 1,
            },
        )
        self.assertEqual(incomplete.status_code, 409)
        compared = self.client.post(
            f"/v1/runs/{run_id}/comparisons/{comparison_id}/actual",
            headers=self.headers,
            json={
                "rows": [
                    {"Mã": "A", "Giá trị": "1.000 ₫"},
                    {"Mã": "B", "Giá trị": "2.000 ₫"},
                ],
                "complete": True,
                "page_count": 3,
                "evidence": "All 3 grid pages and virtual rows inspected.",
            },
        )
        self.assertEqual(compared.status_code, 200, compared.text)
        self.assertEqual(compared.json()["summary"]["matched"], 2)
        self.client.post(
            f"/v1/runs/{run_id}/steps/{step_id}/record",
            headers=self.headers,
            json={
                "status": "passed",
                "actual": "2 rows matched",
                "expected": "2 rows compared",
                "evidence": "All pagination complete",
                "url": "https://sit.hawee.hicas.vn/grid",
            },
        )
        completed = self.client.post(
            f"/v1/runs/{run_id}/complete",
            headers=self.headers,
            json={"status": "completed", "summary": "Comparison completed."},
        )
        self.assertEqual(completed.status_code, 200, completed.text)
        report = self.client.get(
            f"/v1/runs/{run_id}/report?artifact=xlsx",
            headers=self.headers,
        )
        report_path = self.root / "comparison-report.xlsx"
        report_path.write_bytes(report.content)
        executed = load_workbook(report_path, data_only=False)
        self.assertIn("Agent_Data_Comparison", executed.sheetnames)
        self.assertIn("Agent_Run_Log", executed.sheetnames)
        self.assertIn("Notes", executed.sheetnames)
        self.assertEqual(executed["Notes"]["D1"].value, "=SUM(1,2)")
        executed.close()
        self.assertEqual(hashlib.sha256(workbook_path.read_bytes()).hexdigest(), source_hash)

    def test_migration_adds_attention_and_new_tables_without_dropping_run_table(self) -> None:
        database = self.root / "legacy.sqlite3"
        connection = sqlite3.connect(database)
        try:
            connection.execute(
                """
                CREATE TABLE runs (
                    id TEXT PRIMARY KEY,
                    status TEXT NOT NULL,
                    workbook_name TEXT NOT NULL,
                    source_sha256 TEXT NOT NULL,
                    source_path TEXT NOT NULL,
                    output_xlsx_path TEXT,
                    output_html_path TEXT,
                    current_step_id TEXT,
                    summary TEXT NOT NULL DEFAULT '',
                    plan_json TEXT NOT NULL,
                    approval_token_hash TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
            connection.commit()
        finally:
            connection.close()
        QcRepository(database)
        connection = sqlite3.connect(database)
        try:
            columns = {
                row[1] for row in connection.execute("PRAGMA table_info(runs)")
            }
            tables = {
                row[0]
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
        finally:
            connection.close()
        self.assertIn("attention_reason", columns)
        for table in [
            "runs",
            "run_templates",
            "schedules",
            "comparison_results",
            "bug_drafts",
        ]:
            self.assertIn(table, tables)


if __name__ == "__main__":
    unittest.main()
