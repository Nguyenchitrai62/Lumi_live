from __future__ import annotations

import hashlib
import tempfile
import unittest
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from lumi_qc.app import create_app


class QcApiTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.token = "test-installation-token-with-enough-entropy"
        self.client = TestClient(create_app(self.root / "data", self.token))
        self.headers = {
            "Origin": "chrome-extension://abcdefghijklmnop",
            "X-Lumi-Installation-Token": self.token,
        }
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Cases"
        sheet.append(
            [
                "Test Case ID",
                "Title",
                "Test Steps",
                "Expected Result",
                "Actual Result",
                "Status",
            ]
        )
        sheet.append(["TC-1", "Open customer", "Click Customer", "Customer page", "", ""])
        self.workbook = self.root / "input.xlsx"
        workbook.save(self.workbook)
        self.source_hash = hashlib.sha256(self.workbook.read_bytes()).hexdigest()

    def tearDown(self) -> None:
        self.client.close()
        self.temporary.cleanup()

    def test_compile_accepts_reference_workbook(self) -> None:
        reference = Workbook()
        sheet = reference.active
        sheet.title = "ERP"
        sheet.append(
            [
                "No",
                "Feature ID",
                "Module",
                "Feature Name",
                "Description",
                "Business Impact",
                "Priority",
            ]
        )
        sheet.append([1, "", "CRM", "Open customer", "", "", "Critical"])
        reference_path = self.root / "reference.xlsx"
        reference.save(reference_path)
        reference.close()

        compiled = self.client.post(
            "/v1/workbooks/compile",
            headers=self.headers,
            files={
                "file": (
                    "input.xlsx",
                    self.workbook.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                "reference_file": (
                    "reference.xlsx",
                    reference_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"allowed_domains": '["erp.example.com"]'},
        )

        self.assertEqual(compiled.status_code, 200, compiled.text)
        references = compiled.json()["plan"]["reference_workbooks"]
        self.assertEqual(references[0]["name"], "reference.xlsx")
        self.assertEqual(references[0]["stats"]["feature_rows"], 1)

    def test_authenticated_run_lifecycle_and_artifacts(self) -> None:
        self.assertEqual(self.client.get("/health").status_code, 200)
        invalid = self.client.post(
            "/v1/workbooks/compile",
            headers=self.headers,
            files={"file": ("invalid.xlsx", b"not-an-xlsx")},
            data={"allowed_domains": '["erp.example.com"]'},
        )
        self.assertEqual(invalid.status_code, 422)
        denied = self.client.post(
            "/v1/workbooks/compile",
            headers={"Origin": "https://evil.example", "X-Lumi-Installation-Token": self.token},
            files={"file": ("input.xlsx", self.workbook.read_bytes())},
            data={"allowed_domains": '["erp.example.com"]'},
        )
        self.assertEqual(denied.status_code, 403)
        compiled = self.client.post(
            "/v1/workbooks/compile",
            headers=self.headers,
            files={
                "file": (
                    "input.xlsx",
                    self.workbook.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"allowed_domains": '["erp.example.com"]'},
        )
        self.assertEqual(compiled.status_code, 200, compiled.text)
        run = compiled.json()
        run_id = run["run_id"]
        step_id = run["plan"]["test_cases"][0]["steps"][0]["id"]

        approved = self.client.post(f"/v1/runs/{run_id}/approve", headers=self.headers)
        self.assertEqual(approved.status_code, 200, approved.text)
        approval_token = approved.json()["approval_token"]
        self.assertTrue(approval_token)
        started = self.client.post(f"/v1/runs/{run_id}/start", headers=self.headers)
        self.assertEqual(started.json()["status"], "running")
        begun = self.client.post(
            f"/v1/runs/{run_id}/steps/{step_id}/begin",
            headers=self.headers,
        )
        self.assertEqual(begun.status_code, 200, begun.text)

        authorization = self.client.post(
            f"/v1/runs/{run_id}/authorize-action",
            headers=self.headers,
            json={
                "step_id": step_id,
                "action": "browser_click",
                "url": "https://erp.example.com/customers",
                "approval_token": approval_token,
                "arguments": {"index": 3, "confirmed": True},
            },
        )
        self.assertEqual(authorization.status_code, 200)
        self.assertTrue(authorization.json()["authorized"])
        outside_domain = self.client.post(
            f"/v1/runs/{run_id}/authorize-action",
            headers=self.headers,
            json={
                "step_id": step_id,
                "action": "browser_click",
                "url": "https://evil.example/customers",
                "approval_token": approval_token,
                "arguments": {"index": 3, "confirmed": True},
            },
        )
        self.assertFalse(outside_domain.json()["authorized"])

        recorded = self.client.post(
            f"/v1/runs/{run_id}/steps/{step_id}/record",
            headers=self.headers,
            json={
                "status": "passed",
                "actual": "Customer page",
                "expected": "Customer page",
                "evidence": "Heading observed",
                "url": "https://erp.example.com/customers",
            },
        )
        self.assertEqual(recorded.status_code, 200, recorded.text)
        audit_event = self.client.post(
            f"/v1/runs/{run_id}/events",
            headers=self.headers,
            json={
                "type": "redaction_check",
                "phase": "RECORD",
                "step_id": step_id,
                "payload": {
                    "apiKey": "must-not-persist",
                    "message": "password=must-not-persist",
                },
            },
        )
        self.assertEqual(audit_event.status_code, 200)
        self.assertEqual(audit_event.json()["payload"]["apiKey"], "[REDACTED]")
        self.assertNotIn("must-not-persist", str(audit_event.json()["payload"]))
        discovery = self.client.post(
            "/v1/discovery/observations",
            headers=self.headers,
            json={
                "url": "https://erp.example.com/customers",
                "title": "Customer list",
                "tool": "browser_get_page_state",
                "arguments": {},
                "observation": {"content": "Customer table"},
            },
        )
        self.assertEqual(discovery.status_code, 200)
        self.assertEqual(discovery.json()["domain"], "erp.example.com")
        completed = self.client.post(
            f"/v1/runs/{run_id}/complete",
            headers=self.headers,
            json={"status": "completed", "summary": "One step passed."},
        )
        self.assertEqual(completed.status_code, 200, completed.text)
        self.assertEqual(completed.json()["status"], "completed")

        excel = self.client.get(
            f"/v1/runs/{run_id}/report?artifact=xlsx",
            headers=self.headers,
        )
        report = self.client.get(
            f"/v1/runs/{run_id}/report?artifact=html",
            headers=self.headers,
        )
        self.assertEqual(excel.status_code, 200)
        self.assertEqual(report.status_code, 200)
        self.assertIn("Lumi QC execution report", report.text)
        output_path = self.root / "executed.xlsx"
        output_path.write_bytes(excel.content)
        output = load_workbook(output_path)
        self.assertIn("Agent_Run_Log", output.sheetnames)
        output.close()
        self.assertEqual(hashlib.sha256(self.workbook.read_bytes()).hexdigest(), self.source_hash)


if __name__ == "__main__":
    unittest.main()
