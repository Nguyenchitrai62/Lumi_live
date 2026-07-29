from __future__ import annotations

import hashlib
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from lumi_qc.compiler import compile_workbook
from lumi_qc.models import ActionName, Assertion, QcStep, StepResult, WritebackTarget
from lumi_qc.reports import generate_artifacts
from lumi_qc.repository import InvalidTransitionError, QcRepository


class QcServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.input_dir = self.root / "inputs"
        self.output_dir = self.root / "outputs"
        self.input_dir.mkdir()
        self.output_dir.mkdir()
        self.workbook_path = self.root / "cases.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Login"
        sheet.merge_cells("A1:H1")
        sheet["A1"] = "ERP smoke tests"
        sheet.append(
            [
                "Test Case ID",
                "Title",
                "Flow",
                "Test Steps",
                "Test Data",
                "Expected Result",
                "Actual Result",
                "Status",
            ]
        )
        sheet.append(
            [
                "TC-01",
                "Create customer",
                "CRM",
                "Click Create customer",
                "",
                "Create form is visible",
                "",
                "",
            ]
        )
        sheet.append(
            [
                "",
                "",
                "",
                "Enter generated customer name",
                "Lumi 001",
                "",
                "",
                "",
            ]
        )
        sheet["J2"] = "=SUM(1,2)"
        workbook.save(self.workbook_path)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_compile_preserves_source_and_flags_missing_expected(self) -> None:
        original_hash = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        plan, persisted = compile_workbook(
            self.workbook_path,
            self.input_dir,
            ["erp.example.com"],
        )
        self.assertEqual(plan.stats["test_cases"], 1)
        self.assertEqual(plan.stats["steps"], 2)
        self.assertEqual(plan.stats["needs_review"], 1)
        self.assertEqual(plan.stats["spec_issue"], 1)
        self.assertEqual(plan.test_cases[0].steps[0].action.value, "click")
        self.assertEqual(plan.test_cases[0].steps[1].status.value, "spec_issue")
        self.assertEqual(hashlib.sha256(self.workbook_path.read_bytes()).hexdigest(), original_hash)
        self.assertTrue(persisted.exists())

    def test_repository_lifecycle_and_artifacts(self) -> None:
        workbook = load_workbook(self.workbook_path)
        workbook["Login"]["F4"] = "Generated customer name is visible"
        workbook.save(self.workbook_path)
        workbook.close()
        plan, persisted = compile_workbook(
            self.workbook_path,
            self.input_dir,
            ["erp.example.com"],
        )
        repository = QcRepository(self.root / "test.sqlite3")
        repository.create_run(plan, persisted)
        _, approval_token = repository.transition(plan.run_id, "approve")
        self.assertTrue(approval_token)
        repository.transition(plan.run_id, "start")
        step = plan.test_cases[0].steps[0]
        repository.begin_step(plan.run_id, step.id)
        authorization = repository.authorize_action(
            plan.run_id,
            step.id,
            "browser_click",
            "https://erp.example.com/customers",
            approval_token or "",
        )
        self.assertTrue(authorization["authorized"])
        repository.append_event(
            plan.run_id,
            "browser_action_completed",
            "STABILIZE",
            {
                "tool": "browser_click",
                "url": "https://erp.example.com/customers/new",
                "result": {"controllerVerification": {"conclusive": True}},
            },
            step_id=step.id,
        )
        duplicate = repository.authorize_action(
            plan.run_id,
            step.id,
            "browser_click",
            "https://erp.example.com/customers",
            approval_token or "",
        )
        self.assertEqual(duplicate["reason"], "post_action_verification_required")
        repository.record_step(
            plan.run_id,
            step.id,
            StepResult(
                status="passed",
                actual='=HYPERLINK("https://evil.example","evidence")',
                expected=step.expected,
                evidence="Heading observed",
                url="https://erp.example.com/customers/new",
            ),
        )
        original_hash = hashlib.sha256(persisted.read_bytes()).hexdigest()
        xlsx_path, html_path = generate_artifacts(
            plan,
            persisted,
            repository.list_step_results(plan.run_id),
            self.output_dir,
        )
        self.assertEqual(hashlib.sha256(persisted.read_bytes()).hexdigest(), original_hash)
        self.assertTrue(html_path.exists())
        output = load_workbook(xlsx_path, data_only=False)
        self.assertIn("Agent_Run_Log", output.sheetnames)
        self.assertIn("Agent_Result", output.sheetnames)
        self.assertEqual(output["Login"]["H3"].value, "passed")
        self.assertTrue(output["Login"]["G3"].value.startswith("'=HYPERLINK"))
        self.assertEqual(output["Login"]["J2"].value, "=SUM(1,2)")
        output.close()

    def test_high_risk_action_requires_separate_step_approval(self) -> None:
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["Login"]
        sheet["D3"] = "Delete account permanently"
        sheet["F3"] = "Account is deleted"
        sheet["F4"] = "Generated customer name is visible"
        workbook.save(self.workbook_path)
        workbook.close()
        plan, persisted = compile_workbook(
            self.workbook_path,
            self.input_dir,
            ["erp.example.com"],
        )
        repository = QcRepository(self.root / "risk.sqlite3")
        repository.create_run(plan, persisted)
        _, approval_token = repository.transition(plan.run_id, "approve")
        repository.transition(plan.run_id, "start")
        step = plan.test_cases[0].steps[0]
        self.assertEqual(step.risk, "high")
        begun = repository.begin_step(plan.run_id, step.id)
        self.assertTrue(begun["requires_user_approval"])
        denied = repository.authorize_action(
            plan.run_id,
            step.id,
            "browser_click",
            "https://erp.example.com/account",
            approval_token or "",
        )
        self.assertEqual(denied["reason"], "critical_approval_required")
        repository.approve_critical_step(plan.run_id, step.id)
        allowed = repository.authorize_action(
            plan.run_id,
            step.id,
            "browser_click",
            "https://erp.example.com/account",
            approval_token or "",
        )
        self.assertTrue(allowed["authorized"])

    def test_scenario_catalog_uses_domain_adapter_and_reference_catalog(self) -> None:
        scenario_path = self.root / "Business Scenario Catalog.xlsx"
        scenario = Workbook()
        sheet = scenario.active
        sheet.title = "ERP"
        sheet.append(
            [
                "Scenario ID",
                "Nhóm nghiệp vụ",
                "Tên Business Scenario",
                "Chuỗi nghiệp vụ",
                "Type",
                "Notes",
            ]
        )
        sheet.append(
            [
                "E2E-01",
                "Quản trị → Dự án",
                "Khởi tạo dự án",
                "Đăng nhập hệ thống → Tạo dự án",
                "Happy Path",
                "Dự án mới của run",
            ]
        )
        scenario.save(scenario_path)
        scenario.close()

        reference_path = self.root / "Feature_ list.xlsx"
        reference = Workbook()
        feature_sheet = reference.active
        feature_sheet.title = "ERP"
        feature_sheet.append(
            [
                "No",
                "Feature ID",
                "Module",
                "Feature Name",
                "Description",
                "Business Impact",
                "Priority",
            ]
        )
        feature_sheet.append([1, "", "Dự án", "Tạo dự án", "", "", "Critical"])
        detail_sheet = reference.create_sheet("Dự án")
        detail_sheet.append(["No", "Field", "Lable", "Round 1", "Validation"])
        detail_sheet.append([1, "name", "Tên dự án", True, "Bắt buộc"])
        reference.save(reference_path)
        reference.close()

        plan, _ = compile_workbook(
            scenario_path,
            self.input_dir,
            ["sit.hawee.hicas.vn"],
            reference_path=reference_path,
            reference_name=reference_path.name,
        )

        self.assertEqual(plan.stats["test_cases"], 1)
        self.assertEqual(plan.stats["steps"], 4)
        self.assertEqual(plan.stats["needs_review"], 1)
        self.assertEqual(plan.entity_policy.domain, "sit.hawee.hicas.vn")
        self.assertEqual(len(plan.entity_policy.project_markers), 1)
        self.assertEqual(len(plan.generated_data), 1)
        save_step = plan.test_cases[0].steps[-1]
        self.assertEqual(save_step.target, "#button-save-create-update-project-page")
        self.assertEqual(save_step.entity_scope, "project_create")
        self.assertEqual(save_step.status.value, "pending")
        self.assertEqual(plan.reference_workbooks[0].stats["feature_rows"], 1)
        self.assertEqual(plan.reference_workbooks[0].stats["field_spec_rows"], 1)

        repository = QcRepository(self.root / "scenario.sqlite3")
        repository.create_run(plan, self.input_dir / f"{plan.run_id}_{scenario_path.name}")
        with self.assertRaises(InvalidTransitionError):
            repository.transition(plan.run_id, "approve")

    def test_owned_project_gate_requires_verified_run_marker(self) -> None:
        scenario_path = self.root / "owned-project.xlsx"
        scenario = Workbook()
        sheet = scenario.active
        sheet.title = "ERP"
        sheet.append(
            [
                "Scenario ID",
                "Tên Business Scenario",
                "Chuỗi nghiệp vụ",
                "Expected Result",
            ]
        )
        sheet.append(["E2E-OWNED", "Dự án của run", "Tạo dự án", ""])
        scenario.save(scenario_path)
        scenario.close()
        plan, persisted = compile_workbook(
            scenario_path,
            self.input_dir,
            ["sit.hawee.hicas.vn"],
        )
        marker = plan.entity_policy.project_markers[0]
        owned_step = QcStep(
            id="ERP:owned-update",
            order=4,
            instruction="Update the run-owned project",
            action=ActionName.CLICK,
            target="#button-save-owned-project",
            expected="Owned project update saved",
            assertions=[
                Assertion(
                    kind="toast",
                    operator="contains",
                    expected="saved",
                )
            ],
            writeback=WritebackTarget(sheet="ERP"),
            risk="ordinary",
            entity_scope="owned_project",
            source_row=2,
        )
        plan.test_cases[0].steps.append(owned_step)
        plan.stats["steps"] = 4

        repository = QcRepository(self.root / "owned.sqlite3")
        repository.create_run(plan, persisted)
        _, approval_token = repository.transition(plan.run_id, "approve")
        repository.transition(plan.run_id, "start")
        create_steps = plan.test_cases[0].steps[:3]

        repository.begin_step(plan.run_id, create_steps[0].id)
        nav_auth = repository.authorize_action(
            plan.run_id,
            create_steps[0].id,
            "browser_open_tab",
            "https://sit.hawee.hicas.vn/tong-quan?tab=du-an",
            approval_token or "",
            {"url": "https://sit.hawee.hicas.vn/du-an/them"},
        )
        self.assertTrue(nav_auth["authorized"])
        repository.record_step(
            plan.run_id,
            create_steps[0].id,
            StepResult(
                status="passed",
                actual="Tạo dự án",
                expected=create_steps[0].expected,
                evidence="/du-an/them",
            ),
        )

        repository.begin_step(plan.run_id, create_steps[1].id)
        fill_auth = repository.authorize_action(
            plan.run_id,
            create_steps[1].id,
            "browser_input_text",
            "https://sit.hawee.hicas.vn/du-an/them",
            approval_token or "",
            {"text": marker},
        )
        self.assertTrue(fill_auth["authorized"])
        repository.record_step(
            plan.run_id,
            create_steps[1].id,
            StepResult(
                status="passed",
                actual=marker,
                expected=marker,
                evidence=marker,
            ),
        )

        repository.begin_step(plan.run_id, create_steps[2].id)
        repository.append_event(
            plan.run_id,
            "browser_observation_completed",
            "OBSERVE",
            {"result": {"content": f"Tên dự án {marker}"}},
            step_id=create_steps[2].id,
        )
        save_auth = repository.authorize_action(
            plan.run_id,
            create_steps[2].id,
            "browser_click",
            "https://sit.hawee.hicas.vn/du-an/them",
            approval_token or "",
            {"index": 7},
        )
        self.assertTrue(save_auth["authorized"])
        project_id = "34550a5f-326a-3359-8be8-3a22b68ea506"
        repository.append_event(
            plan.run_id,
            "browser_action_completed",
            "STABILIZE",
            {
                "result": {
                    "controllerVerification": {
                        "conclusive": True,
                        "content": (
                            f"{marker} "
                            f"div-project-card-{project_id}-projects-grid-view"
                        ),
                    }
                }
            },
            step_id=create_steps[2].id,
        )
        repository.record_step(
            plan.run_id,
            create_steps[2].id,
            StepResult(
                status="passed",
                actual=marker,
                expected=marker,
                evidence=f"Verified card for {marker}",
            ),
        )
        owned_projects = repository.get_run(plan.run_id)["owned_projects"]
        self.assertEqual(owned_projects[0]["marker"], marker)
        self.assertEqual(owned_projects[0]["project_id"], project_id)

        repository.begin_step(plan.run_id, owned_step.id)
        repository.append_event(
            plan.run_id,
            "browser_observation_completed",
            "OBSERVE",
            {"result": {"content": "Some pre-existing project"}},
            step_id=owned_step.id,
        )
        denied = repository.authorize_action(
            plan.run_id,
            owned_step.id,
            "browser_click",
            f"https://sit.hawee.hicas.vn/du-an/{project_id}",
            approval_token or "",
            {"index": 12},
        )
        self.assertEqual(denied["reason"], "owned_project_evidence_required")
        repository.append_event(
            plan.run_id,
            "browser_observation_completed",
            "OBSERVE",
            {"result": {"content": f"Selected project {marker}"}},
            step_id=owned_step.id,
        )
        allowed = repository.authorize_action(
            plan.run_id,
            owned_step.id,
            "browser_click",
            f"https://sit.hawee.hicas.vn/du-an/{project_id}",
            approval_token or "",
            {"index": 12},
        )
        self.assertTrue(allowed["authorized"])


if __name__ == "__main__":
    unittest.main()
